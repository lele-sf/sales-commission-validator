import os
from openpyxl import load_workbook, Workbook


def parse_value(value):
    try:
        if isinstance(value, str):
            if "R$" in value:
                return float(value.replace("R$", "").replace(".", "").replace(",", "."))
            else:
                return float(value.replace(",", "."))
        return value
    except ValueError as e:
        print(f"Error parsing value: {e}")
        return 0


def convert_to_dict(data, headers):
    return [dict(zip(headers, item)) for item in data]


def load_data(filename, sheet_name):
    try:
        if not filename.lower().endswith(('.xlsx', '.xls')):
            raise ValueError("Invalid file type. Only Excel files are supported.")

        workbook = load_workbook(filename=filename)
        sheet = workbook[sheet_name]
        data = []

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                record = {headers[i]: row[i] for i in range(len(headers))}
                data.append(record)

        return data
    except Exception as e:
        print(f"Error loading data from {filename}: {e}")
        return []


def save_results(data, headers, sheet_name, filename, folder="results"):
    file_path = os.path.join(folder, filename)
    try:
        if not filename.lower().endswith(('.xlsx', '.xls')):
            raise ValueError("Invalid file type. Only Excel files are supported.")

        if not os.path.exists(folder):
            os.makedirs(folder)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        sheet.append(headers)

        for record in data:
            sheet.append([record[header] for header in headers])

        workbook.save(file_path)
    except Exception as e:
        print(f"Error saving results to {file_path}: {e}")
        raise

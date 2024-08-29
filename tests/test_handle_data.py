from openpyxl import Workbook, load_workbook
import pytest
from core.handle_data import convert_to_dict, load_data, parse_value, save_results


def test_parse_value():
    assert parse_value("R$ 1.234,56") == 1234.56
    assert parse_value("1234,56") == 1234.56
    assert parse_value("1234.56") == 1234.56
    assert parse_value(1234.56) == 1234.56
    assert parse_value("invalid") == 0


def test_convert_to_dict():
    data = [
        ("João Silva", 20),
        ("Maria Souza", 15),
        ("Marina Oliveira", 1)
    ]
    headers = ["Nome", "Cotas"]
    expected_output = [
        {"Nome": "João Silva", "Cotas": 20},
        {"Nome": "Maria Souza", "Cotas": 15},
        {"Nome": "Marina Oliveira", "Cotas": 1}
    ]
    assert convert_to_dict(data, headers) == expected_output


def test_load_data_with_valid_file(tmp_path):
    filename = tmp_path / "test_data.xlsx"
    sheet_name = "Sheet1"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    headers = ["Nome do Vendedor", "Valor da Venda"]
    data = [
        ["João Silva", "R$ 1000,00"],
        ["Maria Oliveira", "R$ 500,00"]
    ]

    sheet.append(headers)
    for row in data:
        sheet.append(row)

    workbook.save(filename)

    expected_output = [
        {"Nome do Vendedor": "João Silva", "Valor da Venda": "R$ 1000,00"},
        {"Nome do Vendedor": "Maria Oliveira", "Valor da Venda": "R$ 500,00"}
    ]

    results = load_data(str(filename), sheet_name)
    assert results == expected_output


def test_load_data_invalid_file():
    assert load_data("invalid_file.txt", "Sheet1") == []


def test_save_results(tmp_path):
    file = tmp_path / "test_output.xlsx"
    headers = ["Nome do Vendedor", "Valor da Venda"]
    data = [
        {"Nome do Vendedor": "João Silva", "Valor da Venda": "R$ 1000,00"},
        {"Nome do Vendedor": "Maria Oliveira", "Valor da Venda": "R$ 500,00"}
    ]
    save_results(data, headers, "Sheet1", str(file))

    workbook = load_workbook(filename=str(file))
    sheet = workbook["Sheet1"]
    loaded_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        loaded_data.append(dict(zip(headers, row)))

    expected_output = [
        {"Nome do Vendedor": "João Silva", "Valor da Venda": "R$ 1000,00"},
        {"Nome do Vendedor": "Maria Oliveira", "Valor da Venda": "R$ 500,00"}
    ]
    assert loaded_data == expected_output


def test_save_results_invalid_file():
    headers = ["Nome do Vendedor", "Valor da Venda"]
    data = [
        {"Nome do Vendedor": "João Silva", "Valor da Venda": "R$ 1000,00"},
        {"Nome do Vendedor": "Maria Oliveira", "Valor da Venda": "R$ 500,00"}
    ]
    with pytest.raises(ValueError):
        save_results(data, headers, "Sheet1", "invalid_file.txt")
